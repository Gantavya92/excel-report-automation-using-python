"""
Automated Sales MIS Report Generator
--------------------------------------
Works on real AdventureWorks-style order data (from Microsoft's public
training dataset). Ingests raw order-line data, cleans and reshapes it,
then generates a formatted, formula-driven Excel report.

Usage:
    python report_generator.py raw_sales_data.csv output_report.xlsx
"""

import sys
import re
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.marker import Marker
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice
from openpyxl.formatting.rule import CellIsRule


# ---------- 1. INGEST & CLEAN ----------------------------------------------

def clean_data(path):
    df = pd.read_csv(path)
    before = len(df)

    # The raw "Item" field mixes Product name and Size, e.g. "Mountain-100 Silver, 44"
    # Split it into two real, usable columns.
    split = df["Item"].str.rsplit(",", n=1, expand=True)
    df["Product"] = split[0].str.strip()
    df["Size"] = split[1].str.strip()

    # Pull out just the product line (Mountain-100, Road-650, etc.) for
    # cleaner grouping — the color/size varies but the product line is
    # what a manager actually wants totals by.
    df["Product_Line"] = df["Product"].str.extract(r"^([A-Za-z]+-\d+)")

    # Fix encoding artifacts common in real exports (mangled accented characters)
    df["CustomerName"] = (
        df["CustomerName"]
        .str.replace("�", "n", regex=False)   # common mojibake fallback
    )

    # Revenue doesn't exist in the raw feed — it's a derived MIS metric
    df["Revenue"] = (df["Quantity"] * df["UnitPrice"]).round(2)

    # Parse dates and derive a Month period for trend reporting
    df["OrderDate"] = pd.to_datetime(df["OrderDate"])
    df["Month"] = df["OrderDate"].dt.to_period("M").astype(str)

    # Drop any exact duplicate order lines (defensive — none expected here,
    # but a real recurring pipeline must never assume the next month's file
    # will be as clean as this one)
    duplicates_removed = df.duplicated(subset=["SalesOrderNumber", "SalesOrderLineNumber"]).sum()
    df = df.drop_duplicates(subset=["SalesOrderNumber", "SalesOrderLineNumber"])

    after = len(df)
    log = {
        "rows_in": before,
        "rows_out": after,
        "duplicates_removed": int(duplicates_removed),
        "date_range": f"{df['OrderDate'].min():%d-%b-%Y} to {df['OrderDate'].max():%d-%b-%Y}",
        "distinct_customers": df["CustomerName"].nunique(),
        "distinct_products": df["Product_Line"].nunique(),
        "distinct_orders": df["SalesOrderNumber"].nunique(),
    }
    return df, log


# ---------- 2. AGGREGATE FOR REPORT -----------------------------------------

def build_summaries(df):
    by_product = (
        df.groupby("Product_Line")
        .agg(Units_Sold=("Quantity", "sum"), Revenue=("Revenue", "sum"))
        .reset_index()
        .sort_values("Revenue", ascending=False)
    )
    by_month = (
        df.groupby("Month")
        .agg(Units_Sold=("Quantity", "sum"), Revenue=("Revenue", "sum"))
        .reset_index()
        .sort_values("Month")
    )
    top_customers = (
        df.groupby("CustomerName")
        .agg(Orders=("SalesOrderNumber", "nunique"), Revenue=("Revenue", "sum"))
        .reset_index()
        .sort_values("Revenue", ascending=False)
        .head(10)
    )
    return by_product, by_month, top_customers


# ---------- 3. STYLING HELPERS ----------------------------------------------

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="Arial", bold=True, size=16, color="1F4E78")
LABEL_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CURRENCY_FMT = '"$"#,##0'


def style_header_row(ws, row, n_cols):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
        cell.border = BORDER


def write_table(ws, start_row, start_col, headers, data_rows):
    for j, h in enumerate(headers):
        ws.cell(row=start_row, column=start_col + j, value=h)
    style_header_row(ws, start_row, len(headers))

    for i, row_vals in enumerate(data_rows, start=1):
        for j, val in enumerate(row_vals):
            cell = ws.cell(row=start_row + i, column=start_col + j, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
            if isinstance(val, (int, float)) and j > 0:
                cell.number_format = CURRENCY_FMT if "Revenue" in headers[j] else "#,##0"
    return start_row + len(data_rows) + 1


# ---------- 4. BUILD THE WORKBOOK -------------------------------------------

def build_report(df, log, by_product, by_month, top_customers, out_path):
    wb = Workbook()

    # ---- Sheet 1: Summary dashboard ----
    ws = wb.active
    ws.title = "Summary"

    ws["B2"] = "Sales MIS Report — AdventureWorks Orders"
    ws["B2"].font = TITLE_FONT
    ws["B3"] = f"Generated: {datetime.now():%d-%b-%Y %H:%M}  |  Period covered: {log['date_range']}"
    ws["B3"].font = Font(name="Arial", italic=True, size=9, color="666666")

    last_row = len(df) + 1

    ws["B5"] = "Total Revenue"
    ws["B5"].font = LABEL_FONT
    ws["B6"] = f"=SUM('Raw Data'!I2:I{last_row})"
    ws["B6"].number_format = CURRENCY_FMT
    ws["B6"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")

    ws["D5"] = "Total Units Sold"
    ws["D5"].font = LABEL_FONT
    ws["D6"] = f"=SUM('Raw Data'!G2:G{last_row})"
    ws["D6"].number_format = "#,##0"
    ws["D6"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")

    ws["F5"] = "Number of Orders"
    ws["F5"].font = LABEL_FONT
    # Distinct order count computed once in Python and written as a value,
    # not a live formula — a SUMPRODUCT/COUNTIF distinct-count formula would
    # run an O(n^2) comparison across 30,000+ rows and could hang Excel.
    ws["F6"] = log["distinct_orders"]
    ws["F6"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")
    ws["F6"].number_format = "#,##0"

    ws["H5"] = "Distinct Customers"
    ws["H5"].font = LABEL_FONT
    ws["H6"] = log["distinct_customers"]
    ws["H6"].font = Font(name="Arial", bold=True, size=14, color="1F4E78")

    # Product line summary
    ws["B9"] = "Revenue by Product Line"
    ws["B9"].font = LABEL_FONT
    product_rows = list(by_product.itertuples(index=False, name=None))
    end_row = write_table(ws, 10, 2, ["Product_Line", "Units_Sold", "Revenue"], product_rows)

    ws.conditional_formatting.add(
        f"D11:D{10+len(product_rows)}",
        CellIsRule(operator="lessThan", formula=[f"AVERAGE($D$11:$D${10+len(product_rows)})"],
                   fill=PatternFill("solid", fgColor="FFC7CE")),
    )

    # Top 10 customers
    ws["F9"] = "Top 10 Customers by Revenue"
    ws["F9"].font = LABEL_FONT
    cust_rows = list(top_customers.itertuples(index=False, name=None))
    write_table(ws, 10, 6, ["Customer", "Orders", "Revenue"], cust_rows)

    # Monthly trend table
    month_start = end_row + 2
    ws.cell(row=month_start, column=2, value="Revenue by Month").font = LABEL_FONT
    month_rows = list(by_month.itertuples(index=False, name=None))
    write_table(ws, month_start + 1, 2, ["Month", "Units_Sold", "Revenue"], month_rows)

    # Chart 1: revenue by product line
    chart1 = BarChart()
    chart1.title = "Revenue by Product Line"
    chart1.y_axis.title = "Revenue ($)"
    chart1.style = 10
    data1 = Reference(ws, min_col=4, min_row=10, max_row=10 + len(product_rows))
    cats1 = Reference(ws, min_col=2, min_row=11, max_row=10 + len(product_rows))
    chart1.add_data(data1, titles_from_data=True)
    chart1.set_categories(cats1)
    # Excel's default style auto-colors every bar of a single-series chart
    # and then pads the legend with one entry per category to explain the
    # colors. Force one uniform bar color and drop that redundant legend —
    # the category names already sit on the x-axis.
    chart1.series[0].graphicalProperties.solidFill = "1F4E78"
    chart1.legend = None
    chart1.width = 14
    chart1.height = 8
    ws.add_chart(chart1, f"F{month_start + 1}")

    # Chart 2: revenue trend by month
    chart2 = LineChart()
    chart2.title = "Revenue Trend by Month"
    chart2.y_axis.title = "Revenue ($)"
    chart2.style = 12
    data2 = Reference(ws, min_col=4, min_row=month_start + 1, max_row=month_start + 1 + len(month_rows))
    cats2 = Reference(ws, min_col=2, min_row=month_start + 2, max_row=month_start + 1 + len(month_rows))
    chart2.add_data(data2, titles_from_data=True)
    chart2.set_categories(cats2)
    # Same fix as chart1: one series, one color, no redundant legend.
    # Also disable point markers — with 30 months of data, per-point
    # marker clutter (X shapes) obscures the trend line itself.
    chart2.series[0].graphicalProperties.line.solidFill = "1F4E78"
    chart2.series[0].graphicalProperties.line.width = 20000  # EMUs, ~1.5pt
    chart2.series[0].marker = Marker(symbol="none")
    chart2.series[0].smooth = False
    chart2.legend = None
    chart2.width = 14
    chart2.height = 8
    ws.add_chart(chart2, f"F{month_start + 1 + len(month_rows) + 3}")

    for col, w in zip("ABCDEFGHI", [3, 20, 12, 14, 3, 22, 10, 14, 3]):
        ws.column_dimensions[col].width = w

    # ---- Sheet 2: Data Reconciliation Log ----
    ws2 = wb.create_sheet("Reconciliation Log")
    ws2["B2"] = "Data Cleaning & Reconciliation Log"
    ws2["B2"].font = TITLE_FONT
    log_rows = [
        ("Raw order lines received", log["rows_in"]),
        ("Duplicate order lines removed", log["duplicates_removed"]),
        ("Final clean order lines in report", log["rows_out"]),
        ("Distinct customers", log["distinct_customers"]),
        ("Distinct product lines", log["distinct_products"]),
    ]
    write_table(ws2, 4, 2, ["Step", "Count"], log_rows)
    ws2["B11"] = "Notes"
    ws2["B11"].font = LABEL_FONT
    ws2["B12"] = "Revenue column did not exist in the raw feed — derived as Quantity x UnitPrice."
    ws2["B13"] = "Item field combined product name and size in one string — split into Product and Size columns."
    ws2["B14"] = "Product_Line extracted via regex from Product for cleaner grouping (e.g. 'Mountain-100')."
    for r in (12, 13, 14):
        ws2[f"B{r}"].font = Font(name="Arial", size=9, italic=True, color="666666")
    for col, w in zip("ABC", [3, 55, 12]):
        ws2.column_dimensions[col].width = w

    # ---- Sheet 3: Raw (cleaned) Data ----
    ws3 = wb.create_sheet("Raw Data")
    out_cols = ["SalesOrderNumber", "OrderDate", "CustomerName", "Product",
                "Size", "Product_Line", "Quantity", "UnitPrice", "Revenue", "Month"]
    ws3.append(out_cols)
    style_header_row(ws3, 1, len(out_cols))
    for row in df[out_cols].itertuples(index=False, name=None):
        row = list(row)
        row[1] = row[1].strftime("%d-%m-%Y")
        ws3.append(row)
    ws3.column_dimensions["A"].width = 12
    for col, w in zip("BCDEFGHIJ", [12, 20, 22, 8, 14, 10, 10, 12, 10]):
        ws3.column_dimensions[col].width = w
    ws3.freeze_panes = "A2"

    wb.save(out_path)


# ---------- 5. MAIN ----------------------------------------------------------

if __name__ == "__main__":
    raw_path = sys.argv[1] if len(sys.argv) > 1 else "raw_sales_data.csv"
    out_path = sys.argv[2] if len(sys.argv) > 2 else "Sales_MIS_Report.xlsx"

    df, log = clean_data(raw_path)
    by_product, by_month, top_customers = build_summaries(df)
    build_report(df, log, by_product, by_month, top_customers, out_path)

    print(f"Report generated: {out_path}")
    print("Reconciliation log:", log)
